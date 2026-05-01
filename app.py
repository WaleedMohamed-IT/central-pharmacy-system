"""
app.py - تطبيق Flask الرئيسي
"""

from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, make_response
from functools import wraps
import hashlib
from datetime import datetime, timedelta
import json
from io import BytesIO
import database
import pharmacy_users_management
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
import os

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'your-secret-key-here-change-in-production')
app.config['SESSION_COOKIE_SECURE'] = True
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Strict'

# ==================== DECORATORS ====================

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            flash('برجاء تسجيل الدخول أولاً', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session or session.get('role') != 'admin':
            flash('ليس لديك صلاحيات كافية', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def pharmacist_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session or session.get('role') not in ['pharmacist', 'admin']:
            flash('ليس لديك صلاحيات كافية', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

# ==================== CONTEXT PROCESSORS ====================

@app.context_processor
def inject_user():
    return {
        'current_user': session.get('user'),
        'current_role': session.get('role')
    }

# ==================== ERROR HANDLERS ====================

@app.errorhandler(404)
def not_found(error):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    return render_template('500.html'), 500

# ==================== AUTH ROUTES ====================

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        if not username or not password:
            flash('برجاء إدخال بيانات الدخول', 'warning')
            return redirect(url_for('login'))

        if database.verify_password(username, password):
            user = database.get_user(username)
            
            if not user['is_active']:
                flash('حسابك معطل', 'danger')
                return redirect(url_for('login'))

            session['user'] = user['username']
            session['user_id'] = user['id']
            session['role'] = user['role']
            
            # تسجيل نشاط الدخول
            database.log_activity(
                user['id'],
                'login',
                f'تسجيل دخول المستخدم {username}',
                'user',
                user['id'],
                request.remote_addr
            )
            
            flash(f'مرحباً {username}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('بيانات دخول غير صحيحة', 'danger')
            database.log_activity(0, 'failed_login', f'محاولة دخول فاشلة - {username}', 'user', 0, request.remote_addr)

    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    user_id = session.get('user_id')
    username = session.get('user')
    
    database.log_activity(
        user_id,
        'logout',
        f'تسجيل خروج المستخدم {username}',
        'user',
        user_id,
        request.remote_addr
    )
    
    session.clear()
    flash('تم تسجيل الخروج بنجاح', 'success')
    return redirect(url_for('login'))

# ==================== DASHBOARD ROUTE ====================

@app.route('/dashboard')
@login_required
def dashboard():
    stats = database.get_statistics()
    return render_template('dashboard.html', stats=stats)

# ==================== MEDICINES ROUTES ====================

@app.route('/medicines')
@login_required
def medicines():
    medicines_list = database.get_medicines()
    return render_template('medicines.html', medicines=medicines_list)

@app.route('/search-medicines')
@login_required
def search_medicines():
    search_term = request.args.get('q', '').strip()
    if search_term:
        medicines_list = database.search_medicines(search_term)
    else:
        medicines_list = []
    return render_template('medicines.html', medicines=medicines_list, search_term=search_term)

@app.route('/add', methods=['GET', 'POST'])
@login_required
@pharmacist_required
def add():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        scientific_name = request.form.get('scientific_name', '').strip()
        quantity = request.form.get('quantity', '0').strip()
        pharmacy_type = request.form.get('pharmacy_type', '').strip()
        min_quantity = request.form.get('min_quantity', '10').strip()

        if not name or not pharmacy_type:
            flash('برجاء ملء جميع الحقول المطلوبة', 'warning')
            return redirect(url_for('add'))

        try:
            quantity = int(quantity)
            min_quantity = int(min_quantity)
        except ValueError:
            flash('الكمية يجب أن تكون رقم', 'danger')
            return redirect(url_for('add'))

        database.add_medicine(name, scientific_name, quantity, pharmacy_type, min_quantity)
        
        database.log_activity(
            session.get('user_id'),
            'add_medicine',
            f'إضافة دواء جديد: {name}',
            'medicine',
            None,
            request.remote_addr
        )
        
        flash('تم إضافة الدواء بنجاح', 'success')
        return redirect(url_for('medicines'))

    return render_template('add.html')

@app.route('/edit/<int:medicine_id>', methods=['GET', 'POST'])
@login_required
@pharmacist_required
def edit(medicine_id):
    medicine = database.get_medicine(medicine_id)
    if not medicine:
        flash('الدواء غير موجود', 'danger')
        return redirect(url_for('medicines'))

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        scientific_name = request.form.get('scientific_name', '').strip()
        quantity = request.form.get('quantity', '0').strip()
        pharmacy_type = request.form.get('pharmacy_type', '').strip()
        min_quantity = request.form.get('min_quantity', '10').strip()

        if not name or not pharmacy_type:
            flash('برجاء ملء جميع الحقول المطلوبة', 'warning')
            return redirect(url_for('edit', medicine_id=medicine_id))

        try:
            quantity = int(quantity)
            min_quantity = int(min_quantity)
        except ValueError:
            flash('الكمية يجب أن تكون رقم', 'danger')
            return redirect(url_for('edit', medicine_id=medicine_id))

        database.update_medicine(medicine_id, name, scientific_name, quantity, pharmacy_type, min_quantity)
        
        database.log_activity(
            session.get('user_id'),
            'edit_medicine',
            f'تعديل دواء: {name}',
            'medicine',
            medicine_id,
            request.remote_addr
        )
        
        flash('تم تحديث الدواء بنجاح', 'success')
        return redirect(url_for('medicines'))

    return render_template('edit.html', medicine=medicine)

@app.route('/delete/<int:medicine_id>')
@login_required
@pharmacist_required
def delete(medicine_id):
    medicine = database.get_medicine(medicine_id)
    if not medicine:
        flash('الدواء غير موجود', 'danger')
        return redirect(url_for('medicines'))

    database.delete_medicine(medicine_id)
    
    database.log_activity(
        session.get('user_id'),
        'delete_medicine',
        f'حذف دواء: {medicine["name"]}',
        'medicine',
        medicine_id,
        request.remote_addr
    )
    
    flash('تم حذف الدواء بنجاح', 'success')
    return redirect(url_for('medicines'))

# ==================== DOCTOR ORDER ROUTES ====================

@app.route('/doctor-order', methods=['GET', 'POST'])
@login_required
def doctor_order():
    if session.get('role') not in ['doctor', 'admin']:
        flash('ليس لديك صلاحيات لإضافة طلب', 'danger')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        medicine_id = request.form.get('medicine_id', '').strip()
        quantity = request.form.get('quantity', '0').strip()
        department = request.form.get('department', '').strip()

        if not medicine_id or not quantity or not department:
            flash('برجاء ملء جميع الحقول', 'warning')
            return redirect(url_for('doctor_order'))

        try:
            medicine_id = int(medicine_id)
            quantity = int(quantity)
        except ValueError:
            flash('بيانات غير صحيحة', 'danger')
            return redirect(url_for('doctor_order'))

        medicine = database.get_medicine(medicine_id)
        if not medicine:
            flash('الدواء غير موجود', 'danger')
            return redirect(url_for('doctor_order'))

        order_id = database.add_doctor_order(medicine_id, session.get('user_id'), quantity, department)
        
        database.log_activity(
            session.get('user_id'),
            'create_doctor_order',
            f'إنشاء طلب دواء: {medicine["name"]} - الكمية: {quantity}',
            'doctor_order',
            order_id,
            request.remote_addr
        )
        
        flash('تم إنشاء الطلب بنجاح', 'success')
        return redirect(url_for('pharmacy_orders'))

    medicines = database.get_medicines()
    return render_template('doctor_order.html', medicines=medicines)

# ==================== PHARMACY ORDERS ROUTES ====================

@app.route('/pharmacy-orders')
@login_required
def pharmacy_orders():
    orders = database.get_pharmacy_orders()
    return render_template('pharmacy_orders.html', orders=orders)

@app.route('/approve-order/<int:order_id>', methods=['POST'])
@login_required
@pharmacist_required
def approve_order(order_id):
    notes = request.form.get('notes', '').strip()
    
    database.update_pharmacy_order_status(order_id, 'approved', notes)
    
    database.log_activity(
        session.get('user_id'),
        'approve_order',
        f'الموافقة على طلب رقم {order_id}',
        'pharmacy_order',
        order_id,
        request.remote_addr
    )
    
    flash('تم الموافقة على الطلب بنجاح', 'success')
    return redirect(url_for('pharmacy_orders'))

@app.route('/reject-order/<int:order_id>', methods=['POST'])
@login_required
@pharmacist_required
def reject_order(order_id):
    notes = request.form.get('notes', '').strip()
    
    database.update_pharmacy_order_status(order_id, 'rejected', notes)
    
    database.log_activity(
        session.get('user_id'),
        'reject_order',
        f'رفض طلب رقم {order_id}',
        'pharmacy_order',
        order_id,
        request.remote_addr
    )
    
    flash('تم رفض الطلب', 'success')
    return redirect(url_for('pharmacy_orders'))

# ==================== PRESCRIPTION DETAILS (جديد) ====================

@app.route('/dispense-medicine/<int:order_id>', methods=['GET', 'POST'])
@login_required
@pharmacist_required
def dispense_medicine(order_id):
    """صرف الأدوية وتسجيل التفاصيل"""
    order = database.get_pharmacy_orders()
    order = next((o for o in order if o['id'] == order_id), None)
    
    if not order:
        flash('الطلب غير موجود', 'danger')
        return redirect(url_for('pharmacy_orders'))

    if request.method == 'POST':
        nurse_name = request.form.get('nurse_name', '').strip()
        notes = request.form.get('notes', '').strip()

        if not nurse_name:
            flash('برجاء إدخال اسم الممرضة', 'warning')
            return redirect(url_for('dispense_medicine', order_id=order_id))

        # الحصول على تفاصيل الطلب
        doctor_orders = database.get_doctor_orders()
        doctor_order = next((do for do in doctor_orders if do['id'] == order['doctor_order_id']), None)

        if doctor_order:
            # إضافة تفاصيل الصرف
            database.add_prescription_detail(
                order_id,
                session.get('user_id'),
                nurse_name,
                doctor_order['medicine_id'],
                doctor_order['quantity'],
                notes
            )

            database.log_activity(
                session.get('user_id'),
                'dispense_medicine',
                f'صرف دواء للممرضة {nurse_name}',
                'prescription_detail',
                order_id,
                request.remote_addr
            )

            flash('تم تسجيل الصرف بنجاح', 'success')
            return redirect(url_for('print_prescription', order_id=order_id))

    return render_template('dispense_medicine.html', order=order)

# ==================== REPORTS (جديد) ====================

@app.route('/advanced-reports')
@login_required
def advanced_reports():
    """صفحة التقارير المتقدمة"""
    medicines = database.get_medicines()
    
    start_date = request.args.get('start_date', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
    medicine_id = request.args.get('medicine_id', '')
    report_type = request.args.get('report_type', 'all')
    
    reports = []
    
    if report_type in ['medicine', 'all']:
        # تقرير الأدوية المصروفة
        prescription_details = database.get_activity_log_by_date_range(start_date, end_date)
        if medicine_id:
            reports = [r for r in prescription_details if r.get('target_type') == 'prescription_detail']
        else:
            reports = prescription_details

    return render_template('advanced_reports.html', 
                         medicines=medicines,
                         reports=reports,
                         start_date=start_date,
                         end_date=end_date,
                         medicine_id=medicine_id,
                         report_type=report_type)

# ==================== ACTIVITY LOG (جديد) ====================

@app.route('/activity-log')
@login_required
@admin_required
def activity_log():
    """عرض سجل الأنشطة"""
    user_id = request.args.get('user_id', type=int)
    start_date = request.args.get('start_date', (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
    
    if start_date and end_date:
        logs = database.get_activity_log_by_date_range(start_date, end_date, user_id)
    else:
        logs = database.get_activity_log(user_id, 500)
    
    users = database.get_all_users()
    
    return render_template('activity_log.html', 
                         logs=logs,
                         users=users,
                         selected_user=user_id,
                         start_date=start_date,
                         end_date=end_date)

# ==================== PRINT & PDF (جديد) ====================

@app.route('/print-prescription/<int:order_id>')
@login_required
@pharmacist_required
def print_prescription(order_id):
    """طباعة التقرير"""
    prescription_details = database.get_prescription_details(order_id)
    
    if not prescription_details:
        flash('لا توجد تفاصيل صرف لهذا الطلب', 'danger')
        return redirect(url_for('pharmacy_orders'))

    return render_template('print_prescription.html', 
                         order_id=order_id,
                         details=prescription_details)

@app.route('/export-prescription-pdf/<int:order_id>')
@login_required
@pharmacist_required
def export_prescription_pdf(order_id):
    """تصدير التقرير كـ PDF"""
    prescription_details = database.get_prescription_details(order_id)
    
    if not prescription_details:
        flash('لا توجد تفاصيل صرف', 'danger')
        return redirect(url_for('pharmacy_orders'))

    # إنشاء PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    elements = []

    # العنوان
    style = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=style['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a1a2e'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )

    elements.append(Paragraph('تقرير صرف الأدوية', title_style))
    elements.append(Spacer(1, 0.3*inch))

    # معلومات الطلب
    order_info_data = [
        ['رقم الطلب:', str(order_id)],
        ['التاريخ:', datetime.now().strftime('%Y-%m-%d')],
        ['الوقت:', datetime.now().strftime('%H:%M:%S')]
    ]

    order_info_table = Table(order_info_data, colWidths=[2*inch, 2*inch])
    order_info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f4f8')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey)
    ]))

    elements.append(order_info_table)
    elements.append(Spacer(1, 0.3*inch))

    # جدول الأدوية
    medicines_data = [['اسم الدواء', 'الكمية المصروفة', 'الصيدلي', 'الممرضة']]
    
    for detail in prescription_details:
        medicines_data.append([
            detail['medicine_name'],
            str(detail['quantity_dispensed']),
            detail['pharmacist_name'],
            detail['nurse_name']
        ])

    medicines_table = Table(medicines_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    medicines_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(medicines_table)
    elements.append(Spacer(1, 0.5*inch))

    # التواقيع
    signature_style = ParagraphStyle(
        'SignatureStyle',
        parent=style['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        spaceAfter=60
    )

    elements.append(Spacer(1, 0.3*inch))
    elements.append(Paragraph('ملاحظة: يرجى التوقيع أدناه بعد الانتهاء من الصرف', signature_style))

    # جدول التواقيع
    signature_data = [
        ['توقيع الصيدلي', 'التاريخ', 'توقيع الممرضة', 'التاريخ'],
        ['_____________', '___________', '_____________', '___________']
    ]

    signature_table = Table(signature_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    signature_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, 1), 30)
    ]))

    elements.append(signature_table)

    # بناء PDF
    doc.build(elements)
    buffer.seek(0)

    response = make_response(buffer.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename=prescription_{order_id}.pdf'
    response.headers['Content-Type'] = 'application/pdf'

    return response

# ==================== EXPORT ROUTES ====================

@app.route('/export-excel')
@login_required
def export_excel():
    """تصدير البيانات إلى Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الأدوية"

    # رؤوس الأعمدة
    headers = ['رقم الدواء', 'الاسم', 'الاسم العلمي', 'الكمية', 'نوع الصيدلية', 'الكمية الدنيا']
    ws.append(headers)

    # تنسيق الرؤوس
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # إضافة البيانات
    medicines = database.get_medicines()
    for medicine in medicines:
        ws.append([
            medicine['id'],
            medicine['name'],
            medicine['scientific_name'],
            medicine['quantity'],
            medicine['pharmacy_type'],
            medicine['min_quantity']
        ])

    # حفظ الملف
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = make_response(buffer.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=medicines.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return response

@app.route('/export-pdf')
@login_required
def export_pdf():
    """تصدير البيانات إلى PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    style = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        parent=style['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1a1a2e'),
        spaceAfter=30,
        alignment=TA_CENTER
    )

    elements.append(Paragraph('قائمة الأدوية', title_style))
    elements.append(Spacer(1, 0.3*inch))

    # جدول الأدوية
    medicines = database.get_medicines()
    data = [['رقم', 'الدواء', 'الاسم العلمي', 'الكمية', 'النوع']]
    
    for med in medicines:
        data.append([
            str(med['id']),
            med['name'],
            med['scientific_name'],
            str(med['quantity']),
            med['pharmacy_type']
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    response = make_response(buffer.getvalue())
    response.headers['Content-Disposition'] = 'attachment; filename=medicines.pdf'
    response.headers['Content-Type'] = 'application/pdf'

    return response

# ==================== BLUEPRINT: USERS ====================

app.register_blueprint(pharmacy_users_management.users_bp)

# ==================== MAIN ====================

if __name__ == '__main__':
    database.init_db()
    app.run(debug=False, host='0.0.0.0', port=5000)